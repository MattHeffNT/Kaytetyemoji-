import os

import openpyxl

# open the source workbook
source_wb = openpyxl.load_workbook("./_Kaytetyemoji.xlsx", data_only=True)

# get the active sheet in each workbook
source_sheet = source_wb.active
new_sheet = source_wb.create_sheet("New Sheet")

# this path is just for where I've stored the emojis, change according to your file structure
emoji_path = "/home/heffa/Downloads/EmojisAndAudio/emojis/"
audio_path = "/home/heffa/Downloads/EmojisAndAudio/audio/"
phrase_path = "/home/heffa/Downloads/EmojisAndAudio/audio/phrases/"

emoji_list = os.listdir(emoji_path)
audio_list = os.listdir(audio_path)
phrase_list = os.listdir(phrase_path)

# grab the emoji images in the folder and add them to our spreadsheet (new sheet) in filename

# we just need this variable while we wait for the moon so we can skip over missing items
k = 0

# emojis
for i, j in enumerate(sorted(emoji_list), start=1):

    # if i == 49:
    #     new_sheet.cell(row=i + 1, column=1).value = "empty"
    #     k = 1

    new_sheet.cell(row=i + 1 + k, column=1).value = j

# audio
for i, j in enumerate(sorted(audio_list), start=1):

    # ignore phrase folder and just grab audio
    if j.endswith(".mp3"):
        new_sheet.cell(row=i + 1, column=4).value = j

# phrases
for i, j in enumerate(sorted(phrase_list), start=1):

    # if phrase isn't recorded yet, then skip row
    if source_sheet.cell(row=i + 1, column=8).value is None:
        new_sheet.cell(row=i + 1, column=5).value = ""

        continue

    new_sheet.cell(row=i + 1, column=5).value = j


# iterate through the rows and copy the values
for i in range(1, source_sheet.max_row + 1):

    # get the value from column A (Kaytetye name)
    value_a = source_sheet.cell(row=i, column=1).value

    # get the value from column D (english name)
    value_d = source_sheet.cell(row=i, column=4).value

    # get the value from column H (phrase)
    value_h = source_sheet.cell(row=i, column=8).value

    # just until this is done *** Remove once we get the moon
    if value_d == "Moon":
        new_sheet.cell(row=i, column=1).value = ""

    # paste copied values into new columns on new sheet
    # name (english)
    new_sheet.cell(row=i, column=2).value = value_d

    # name_kaytetye
    new_sheet.cell(row=i, column=3).value = value_a

    # phrases
    new_sheet.cell(row=i, column=5).value = value_h

# add column names
new_sheet["a1"] = "filename"
new_sheet["b1"] = "name"
new_sheet["c1"] = "name_kaytetye"
new_sheet["d1"] = "audio"
new_sheet["e1"] = "phrases"

#  save the new workbook
source_wb.save("./_Kaytetyemoji.xlsx")
